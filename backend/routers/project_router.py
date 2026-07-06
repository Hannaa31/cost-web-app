import sys, os, io
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from typing import List, Optional
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from database import get_db
import models, schemas, auth

router = APIRouter(prefix="/api/projects", tags=["Projects & Workspace"])

@router.get("", response_model=List[schemas.ProjectResponse])
def list_user_projects(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    # Admins see all projects or user sees their own projects
    if current_user.role == models.UserRole.admin:
        return db.query(models.Project).order_by(models.Project.created_at.desc()).all()
    return db.query(models.Project).filter(models.Project.created_by == current_user.id).order_by(models.Project.created_at.desc()).all()

@router.post("", response_model=schemas.ProjectResponse, status_code=status.HTTP_201_CREATED)
def create_project(
    project_in: schemas.ProjectCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    new_project = models.Project(
        created_by=current_user.id,
        name=project_in.name,
        client=project_in.client,
        global_margin_pct=project_in.global_margin_pct,
        global_erection_pct=project_in.global_erection_pct,
        default_annual_escalation_pct=project_in.default_annual_escalation_pct
    )
    db.add(new_project)
    db.commit()
    db.refresh(new_project)
    return new_project

@router.get("/{project_id}", response_model=schemas.ProjectResponse)
def get_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != models.UserRole.admin and project.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to access this workspace")
    return project

@router.put("/{project_id}", response_model=schemas.ProjectResponse)
def update_project(
    project_id: int,
    project_update: schemas.ProjectUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != models.UserRole.admin and project.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to modify this workspace")
    
    if project_update.name is not None:
        project.name = project_update.name
    if project_update.client is not None:
        project.client = project_update.client
    if project_update.global_margin_pct is not None:
        project.global_margin_pct = project_update.global_margin_pct
    if project_update.global_erection_pct is not None:
        project.global_erection_pct = project_update.global_erection_pct
    
    if project_update.default_annual_escalation_pct is not None and project_update.default_annual_escalation_pct != project.default_annual_escalation_pct:
        project.default_annual_escalation_pct = project_update.default_annual_escalation_pct
        line_items = db.query(models.EstimateLineItem).filter(models.EstimateLineItem.project_id == project_id).all()
        now = datetime.utcnow()
        for item in line_items:
            rate = db.query(models.MasterRate).filter(models.MasterRate.id == item.selected_rate_id).first()
            if rate:
                years_elapsed = max(0.0, (now - rate.quotation_date).days / 365.25)
                multiplier = (1 + project.default_annual_escalation_pct) ** years_elapsed
                item.calculated_escalated_rate = round(rate.base_rate * multiplier, 2)
                item.total_item_cost = round(item.calculated_escalated_rate * item.quantity, 2)
    elif project_update.default_annual_escalation_pct is not None:
        project.default_annual_escalation_pct = project_update.default_annual_escalation_pct

    db.commit()
    db.refresh(project)
    return project

@router.delete("/{project_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != models.UserRole.admin and project.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this workspace")
    
    # Delete associated line items first
    db.query(models.EstimateLineItem).filter(models.EstimateLineItem.project_id == project_id).delete()
    db.delete(project)
    db.commit()
    return None

@router.get("/{project_id}/line-items", response_model=List[schemas.EstimateLineItemResponse])
def get_project_line_items(
    project_id: int,
    domain: Optional[models.DomainType] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    query = db.query(models.EstimateLineItem).filter(models.EstimateLineItem.project_id == project_id)
    if domain:
        query = query.filter(models.EstimateLineItem.domain == domain)
    return query.all()

@router.post("/{project_id}/line-items", response_model=schemas.EstimateLineItemResponse, status_code=status.HTTP_201_CREATED)
def add_line_item(
    project_id: int,
    item_in: schemas.EstimateLineItemCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    rate = db.query(models.MasterRate).filter(models.MasterRate.id == item_in.selected_rate_id).first()
    if not rate:
        raise HTTPException(status_code=404, detail="Selected Master Rate not found")
    
    cat = db.query(models.EquipmentCategory).filter(models.EquipmentCategory.id == item_in.category_id).first()
    item_domain = item_in.domain or (cat.domain if cat else models.DomainType.Mechanical)

    # Calculate escalated rate using project default escalation
    now = datetime.utcnow()
    years_elapsed = max(0.0, (now - rate.quotation_date).days / 365.25)
    multiplier = (1 + project.default_annual_escalation_pct) ** years_elapsed
    escalated_rate = round(rate.base_rate * multiplier, 2)
    total_cost = round(escalated_rate * item_in.quantity, 2)

    new_item = models.EstimateLineItem(
        project_id=project_id,
        category_id=item_in.category_id,
        selected_rate_id=item_in.selected_rate_id,
        domain=item_domain,
        quantity=item_in.quantity,
        calculated_escalated_rate=escalated_rate,
        total_item_cost=total_cost
    )
    db.add(new_item)
    db.commit()
    db.refresh(new_item)
    return new_item

@router.put("/{project_id}/line-items/{item_id}", response_model=schemas.EstimateLineItemResponse)
def update_line_item(
    project_id: int,
    item_id: int,
    item_update: schemas.EstimateLineItemUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    item = db.query(models.EstimateLineItem).filter(
        models.EstimateLineItem.id == item_id,
        models.EstimateLineItem.project_id == project_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Line item not found")
    
    if item_update.selected_rate_id is not None and item_update.selected_rate_id != item.selected_rate_id:
        new_rate = db.query(models.MasterRate).filter(models.MasterRate.id == item_update.selected_rate_id).first()
        if not new_rate:
            raise HTTPException(status_code=404, detail="New Master Rate not found")
        item.selected_rate_id = new_rate.id
        now = datetime.utcnow()
        years_elapsed = max(0.0, (now - new_rate.quotation_date).days / 365.25)
        multiplier = (1 + project.default_annual_escalation_pct) ** years_elapsed
        item.calculated_escalated_rate = round(new_rate.base_rate * multiplier, 2)
    
    if item_update.quantity is not None:
        item.quantity = item_update.quantity

    item.total_item_cost = round(item.calculated_escalated_rate * item.quantity, 2)
    db.commit()
    db.refresh(item)
    return item

@router.delete("/{project_id}/line-items/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_line_item(
    project_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    item = db.query(models.EstimateLineItem).filter(
        models.EstimateLineItem.id == item_id,
        models.EstimateLineItem.project_id == project_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Line item not found")
    db.delete(item)
    db.commit()
    return None

def populate_sheet(ws, project, line_items, domain_name=None):
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=10, bold=True)
    header_fill = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
    accent_fill = PatternFill(start_color="0284c7" if not domain_name else ("059669" if domain_name == "Mechanical" else ("d97706" if domain_name == "Electrical" else "4f46e5")), end_color="0284c7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    if domain_name:
        ws.title = domain_name
        items = [i for i in line_items if (i.domain == domain_name or (i.category and i.category.domain == domain_name))]
    else:
        ws.title = "Cost Report"
        items = line_items

    ws.merge_cells("A1:I1")
    ws["A1"] = f"PROJECT ESTIMATION COST REPORT: {project.name.upper()} ({domain_name.upper() if domain_name else 'ALL DOMAINS'})"
    ws["A1"].font = title_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Client Entity:"
    ws["B3"] = project.client
    ws["A4"] = "Generated Date:"
    ws["B4"] = datetime.now().strftime("%d/%m/%Y")
    ws["D3"] = "Margin:"
    ws["E3"] = f"{project.global_margin_pct * 100:.1f}%"
    ws["D4"] = "Erection:"
    ws["E4"] = f"{project.global_erection_pct * 100:.1f}%"
    for cell in ["A3", "A4", "D3", "D4"]:
        ws[cell].font = bold_font

    headers = ["Item #", "Equipment Category", "Specifications", "Vendor Name", "Base Rate (INR)", "Escalated Rate (INR)", "Remarks", "Quantity", "Total Cost (INR)"]
    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num, value=h_text)
        cell.font = header_font
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[6].height = 22

    subtotal = 0.0
    row_idx = 7
    for i, item in enumerate(items, 1):
        cat_name = item.category.name if item.category else f"Category #{item.category_id}"
        vendor_name = item.selected_rate.vendor_name if item.selected_rate else f"Rate #{item.selected_rate_id}"
        base_rate = item.selected_rate.base_rate if item.selected_rate else 0.0
        
        spec_dict = item.selected_rate.specifications if item.selected_rate and item.selected_rate.specifications else {}
        if isinstance(spec_dict, dict) and spec_dict:
            spec_str = ", ".join(f"{k}: {v}" for k, v in spec_dict.items() if k != "Remarks")
        else:
            spec_str = str(spec_dict) if spec_dict else "Standard"
        
        remarks_str = ""
        if item.selected_rate:
            remarks_str = item.selected_rate.remarks or (spec_dict.get("Remarks") if isinstance(spec_dict, dict) else "") or "-"

        ws.cell(row=row_idx, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=cat_name)
        ws.cell(row=row_idx, column=3, value=spec_str)
        ws.cell(row=row_idx, column=4, value=vendor_name)
        
        base_cell = ws.cell(row=row_idx, column=5, value=base_rate)
        base_cell.number_format = '₹#,##0.00'
        base_cell.alignment = Alignment(horizontal="right")

        rate_cell = ws.cell(row=row_idx, column=6, value=item.calculated_escalated_rate)
        rate_cell.number_format = '₹#,##0.00'
        rate_cell.alignment = Alignment(horizontal="right")
        
        ws.cell(row=row_idx, column=7, value=remarks_str)

        qty_cell = ws.cell(row=row_idx, column=8, value=item.quantity)
        qty_cell.alignment = Alignment(horizontal="right")
        
        cost_cell = ws.cell(row=row_idx, column=9, value=item.total_item_cost)
        cost_cell.number_format = '₹#,##0.00'
        cost_cell.alignment = Alignment(horizontal="right")
        
        for c in range(1, 10):
            ws.cell(row=row_idx, column=c).border = thin_border
            
        subtotal += item.total_item_cost
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

    if len(items) == 0:
        ws.merge_cells(f"A{row_idx}:I{row_idx}")
        ws.cell(row=row_idx, column=1, value=f"No {domain_name or ''} equipment added yet.").alignment = Alignment(horizontal="center")
        row_idx += 1

    erection_cost = round(subtotal * project.global_erection_pct, 2)
    margin_cost = round(subtotal * project.global_margin_pct, 2)
    grand_total = round(subtotal + erection_cost + margin_cost, 2)

    row_idx += 1
    summary_rows = [
        ("Equipment Subtotal:", subtotal),
        (f"Margin ({project.global_margin_pct * 100:.1f}%):", margin_cost),
        (f"Erection ({project.global_erection_pct * 100:.1f}%):", erection_cost),
        ("GRAND ESTIMATE TOTAL:", grand_total)
    ]

    for label, val in summary_rows:
        ws.merge_cells(f"G{row_idx}:H{row_idx}")
        label_cell = ws.cell(row=row_idx, column=7, value=label)
        label_cell.font = bold_font if label != "GRAND ESTIMATE TOTAL:" else Font(name="Calibri", size=11, bold=True, color="004B23")
        label_cell.alignment = Alignment(horizontal="right")
        
        val_cell = ws.cell(row=row_idx, column=9, value=val)
        val_cell.font = bold_font if label != "GRAND ESTIMATE TOTAL:" else Font(name="Calibri", size=11, bold=True, color="004B23")
        val_cell.number_format = '₹#,##0.00'
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.border = thin_border
        ws.row_dimensions[row_idx].height = 20 if label == "GRAND ESTIMATE TOTAL:" else 18
        row_idx += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
    return subtotal

@router.get("/{project_id}/export-excel")
def export_project_excel(
    project_id: int,
    domain: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    line_items = db.query(models.EstimateLineItem).filter(models.EstimateLineItem.project_id == project_id).all()
    wb = openpyxl.Workbook()

    if domain in ["Mechanical", "Electrical", "Civil"]:
        ws = wb.active
        populate_sheet(ws, project, line_items, domain)
        filename = f"{domain}_Report.xlsx"
    else:
        # Full 4-sheet project report
        ws_mech = wb.active
        mech_sub = populate_sheet(ws_mech, project, line_items, "Mechanical")
        
        ws_elec = wb.create_sheet(title="Electrical")
        elec_sub = populate_sheet(ws_elec, project, line_items, "Electrical")
        
        ws_civil = wb.create_sheet(title="Civil")
        civil_sub = populate_sheet(ws_civil, project, line_items, "Civil")
        
        # Sheet 4: Summary Sheet
        ws_sum = wb.create_sheet(title="Summary")
        title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
        bold_font = Font(name="Calibri", size=11, bold=True)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        ws_sum.merge_cells("A1:F1")
        ws_sum["A1"] = f"EXECUTIVE PROJECT ESTIMATION SUMMARY: {project.name.upper()}"
        ws_sum["A1"].font = title_font
        ws_sum["A1"].fill = header_fill
        ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 35
        
        ws_sum["A3"] = "Client Name:"
        ws_sum["B3"] = project.client
        ws_sum["A4"] = "Date Generated:"
        ws_sum["B4"] = datetime.now().strftime("%d/%m/%Y")
        ws_sum["D3"] = "Global Margin:"
        ws_sum["E3"] = f"{project.global_margin_pct * 100:.1f}%"
        ws_sum["D4"] = "Global Erection:"
        ws_sum["E4"] = f"{project.global_erection_pct * 100:.1f}%"
        for cell in ["A3", "A4", "D3", "D4", "B3", "B4", "E3", "E4"]:
            ws_sum[cell].font = bold_font

        headers = ["Discipline / Category", "Line Items Count", "Total Equipment Cost (INR)"]
        for col_num, h_text in enumerate(headers, 2):
            cell = ws_sum.cell(row=6, column=col_num, value=h_text)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws_sum.row_dimensions[6].height = 24
        
        mech_count = len([i for i in line_items if i.domain == "Mechanical" or (i.category and i.category.domain == "Mechanical")])
        elec_count = len([i for i in line_items if i.domain == "Electrical" or (i.category and i.category.domain == "Electrical")])
        civil_count = len([i for i in line_items if i.domain == "Civil" or (i.category and i.category.domain == "Civil")])
        
        subtotal = mech_sub + elec_sub + civil_sub
        margin_cost = round(subtotal * project.global_margin_pct, 2)
        erection_cost = round(subtotal * project.global_erection_pct, 2)
        grand_total = round(subtotal + margin_cost + erection_cost, 2)
        
        rows_data = [
            ("Mechanical Discipline", mech_count, mech_sub),
            ("Electrical Discipline", elec_count, elec_sub),
            ("Civil Discipline", civil_count, civil_sub),
            ("EQUIPMENT SUBTOTAL", mech_count + elec_count + civil_count, subtotal),
            (f"Margin Amount ({project.global_margin_pct * 100:.1f}%)", "-", margin_cost),
            (f"Erection Amount ({project.global_erection_pct * 100:.1f}%)", "-", erection_cost),
            ("GRAND ESTIMATE TOTAL", "-", grand_total)
        ]
        
        r_idx = 7
        for label, count, val in rows_data:
            ws_sum.cell(row=r_idx, column=2, value=label).font = bold_font
            ws_sum.cell(row=r_idx, column=3, value=count).alignment = Alignment(horizontal="center")
            val_cell = ws_sum.cell(row=r_idx, column=4, value=val)
            val_cell.number_format = '₹#,##0.00'
            val_cell.alignment = Alignment(horizontal="right")
            val_cell.font = bold_font if "TOTAL" in label or "SUBTOTAL" in label else Font(name="Calibri", size=11)
            if "GRAND" in label:
                val_cell.font = Font(name="Calibri", size=13, bold=True, color="004B23")
            for c in range(2, 5):
                ws_sum.cell(row=r_idx, column=c).border = thin_border
            ws_sum.row_dimensions[r_idx].height = 22 if "TOTAL" in label else 19
            r_idx += 1
            
        for col in ws_sum.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 16)
            
        clean_name = "".join(x for x in project.name if x.isalnum() or x in " _-").replace(" ", "_")
        filename = f"{clean_name}_Full_Project_Report.xlsx"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
